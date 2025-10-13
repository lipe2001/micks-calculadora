import os
from typing import Optional, List
from fastapi import FastAPI, Request, Form, status, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import httpx

API_BASE_URL = os.getenv("API_BASE_URL", "http://api:3000")
ADMIN_USER = "admin"
ADMIN_PASS = "desafio"

app = FastAPI(title="Micks Calculadora WEB")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("WEB_SESSION_SECRET", "dev-secret-please-change"))

templates = Jinja2Templates(directory="templates")

@app.get("/health")
def health():
    return {"status": "ok", "service": "web"}

# ---------- HOME -> calculadora ----------
@app.get("/", response_class=HTMLResponse)
def home():
    return RedirectResponse(url="/calculadora_plano", status_code=status.HTTP_302_FOUND)

# ---------- LOGIN / VENDAS ----------
def is_logged(request: Request) -> bool:
    return request.session.get("user") == ADMIN_USER

@app.get("/login", response_class=HTMLResponse)
def get_login(request: Request):
    if is_logged(request):
        return RedirectResponse(url="/vendas", status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})

@app.post("/login", response_class=HTMLResponse)
def post_login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == ADMIN_USER and password == ADMIN_PASS:
        request.session["user"] = ADMIN_USER
        return RedirectResponse(url="/vendas", status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse("login.html", {"request": request, "error": "Usuário ou senha inválidos."})

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)

@app.get("/vendas", response_class=HTMLResponse)
async def vendas(
    request: Request,
    q: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    order_by: str = "created_at",
    order_dir: str = "desc",
):
    if not is_logged(request):
        return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)

    params = {
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "order_by": order_by,
        "order_dir": order_dir,
        "limit": "200",
        "offset": "0",
    }
    async with httpx.AsyncClient(base_url=API_BASE_URL, timeout=10.0) as client:
        r = await client.get(
            "/api/sales",
            params={k: v for k, v in params.items() if v},
            auth=(ADMIN_USER, ADMIN_PASS),
        )
        r.raise_for_status()
        sales = r.json()

    return templates.TemplateResponse(
        "vendas.html",
        {
            "request": request,
            "sales": sales,
            "q": q or "",
            "date_from": date_from or "",
            "date_to": date_to or "",
            "order_by": order_by,
            "order_dir": order_dir,
            "xlsx_rows": None,  # preenchido quando fizer upload (opcional)
        },
    )

# --------- (Opcional) Upload XLSX (preview no admin, não persiste) ----------
@app.post("/vendas/upload", response_class=HTMLResponse)
async def vendas_upload(request: Request, file: UploadFile = File(...)):
    if not is_logged(request):
        return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)

    # Leitura rápida do XLSX
    from openpyxl import load_workbook
    content = await file.read()
    from io import BytesIO
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active  # usa a primeira planilha

    # Coleta até ~200 linhas para preview
    rows: List[List[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 200:
            break
        rows.append([("" if c is None else str(c)) for c in row])

    # Reapresenta a tela de vendas só com o preview (sem reconsultar API)
    return templates.TemplateResponse(
        "partials/xlsx_preview.html",
        {
            "request": request,
            "filename": file.filename,
            "rows": rows,
        },
    )

# ---------- CALCULADORA / CONTRATAR ----------
@app.get("/calculadora_plano", response_class=HTMLResponse)
def calculadora_plano(request: Request):
    return templates.TemplateResponse("calculadora_plano.html", {"request": request})

@app.post("/_calc", response_class=HTMLResponse)
async def _calc(
    request: Request,
    n_cellphones: int = Form(0),
    n_computers: int = Form(0),
    n_smart_tvs: int = Form(0),
    n_tv_box: int = Form(0),
    n_others: int = Form(0),
    gamer: Optional[bool] = Form(False),
):
    payload = {
        "n_cellphones": max(0, int(n_cellphones)),
        "n_computers": max(0, int(n_computers)),
        "n_smart_tvs": max(0, int(n_smart_tvs)),
        "n_tv_box": max(0, int(n_tv_box)),
        "n_others": max(0, int(n_others)),
        "gamer": bool(gamer),
    }
    async with httpx.AsyncClient(base_url=API_BASE_URL, timeout=10.0) as client:
        r = await client.post("/api/calculate", json=payload)
        r.raise_for_status()
        result = r.json()

    ctx = {"request": request, "result": result, "payload": payload}
    return templates.TemplateResponse("partials/calc_result.html", ctx)

@app.post("/_contract", response_class=HTMLResponse)
async def _contract(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    n_cellphones: int = Form(0),
    n_computers: int = Form(0),
    n_smart_tvs: int = Form(0),
    n_tv_box: int = Form(0),
    n_others: int = Form(0),
    gamer: Optional[bool] = Form(False),
):
    payload = {
        "name": name.strip(),
        "email": email.strip(),
        "phone": phone.strip(),
        "n_cellphones": max(0, int(n_cellphones)),
        "n_computers": max(0, int(n_computers)),
        "n_smart_tvs": max(0, int(n_smart_tvs)),
        "n_tv_box": max(0, int(n_tv_box)),
        "n_others": max(0, int(n_others)),
        "gamer": bool(gamer),
    }
    async with httpx.AsyncClient(base_url=API_BASE_URL, timeout=10.0) as client:
        r = await client.post("/api/contract", json=payload)
        r.raise_for_status()
        data = r.json()

    ctx = {"request": request, "data": data, "payload": payload}
    return templates.TemplateResponse("partials/contract_success.html", ctx)
